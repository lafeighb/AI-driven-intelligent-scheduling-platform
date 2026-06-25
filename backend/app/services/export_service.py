"""排课结果导出服务 - 支持Excel和PDF格式"""
import io
from typing import List, Dict, Any, Optional
from datetime import datetime


class ExportService:
    """报表导出服务"""

    @staticmethod
    def export_to_excel(schedule_entries: List[Dict], metadata: Dict = None) -> bytes:
        """导出排课方案到Excel格式

        Args:
            schedule_entries: 排课条目列表
            metadata: 元数据（版本号、统计信息等）

        Returns:
            Excel文件的二进制内容
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # === 工作表1: 按班级查看 ===
        ws_class = wb.active
        ws_class.title = "按班级课表"
        ExportService._build_timetable_sheet(ws_class, schedule_entries, "class_name", "班级课表")

        # === 工作表2: 按教师查看 ===
        ws_teacher = wb.create_sheet("按教师课表")
        ExportService._build_timetable_sheet(ws_teacher, schedule_entries, "teacher_name", "教师课表")

        # === 工作表3: 按教室查看 ===
        ws_room = wb.create_sheet("按教室课表")
        ExportService._build_timetable_sheet(ws_room, schedule_entries, "classroom_name", "教室课表")

        # === 工作表4: 全部排课数据 ===
        ws_all = wb.create_sheet("全部排课数据")
        headers = ["序号", "课程名称", "教师", "班级", "教室", "周次", "星期",
                   "开始节次", "结束节次", "连排", "状态"]

        # 表头样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col, header in enumerate(headers, 1):
            cell = ws_all.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        day_names = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五", 6: "周六", 7: "周日"}

        for i, entry in enumerate(schedule_entries, 2):
            row_data = [
                i - 1,
                entry.get("course_name", ""),
                entry.get("teacher_name", ""),
                entry.get("class_name", ""),
                entry.get("classroom_name", ""),
                f"第{entry.get('week_number', 1)}周",
                day_names.get(entry.get("day_of_week", 1), ""),
                f"第{entry.get('period_start', 1)}节",
                f"第{entry.get('period_end', 1)}节",
                "是" if entry.get("is_consecutive") else "否",
                entry.get("status", "confirmed"),
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws_all.cell(row=i, column=col, value=value)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws_all.column_dimensions[get_column_letter(col)].width = 14

        # === 工作表5: 统计信息 ===
        if metadata:
            ws_stats = wb.create_sheet("统计信息")
            stats_data = [
                ["排课统计报告", "", ""],
                ["", "", ""],
                ["版本号", metadata.get("version", ""), ""],
                ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""],
                ["总排课条目", metadata.get("total_entries", 0), ""],
                ["硬性冲突数", metadata.get("hard_conflicts", 0), ""],
                ["软性冲突数", metadata.get("soft_conflicts", 0), ""],
                ["质量评分", f"{metadata.get('quality_score', 0)}分", ""],
                ["排课算法", metadata.get("algorithm", ""), ""],
                ["执行时间", f"{metadata.get('execution_time', 0)}秒", ""],
            ]
            for i, row in enumerate(stats_data, 1):
                for j, val in enumerate(row, 1):
                    cell = ws_stats.cell(row=i, column=j, value=val)
                    if i == 1:
                        cell.font = Font(bold=True, size=14)

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _build_timetable_sheet(worksheet, entries: List[Dict], group_field: str, title: str):
        """构建课表视图工作表"""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # 分组
        groups: Dict[str, List[Dict]] = {}
        for entry in entries:
            key = entry.get(group_field, "未分类")
            if key not in groups:
                groups[key] = []
            groups[key].append(entry)

        day_names = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五"}
        max_period = 8

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=10, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        current_row = 1
        for group_name, group_entries in sorted(groups.items()):
            # 分组标题
            cell = worksheet.cell(row=current_row, column=1, value=f"{group_name}")
            cell.font = Font(bold=True, size=12)
            current_row += 1

            # 表头
            headers = ["节次"] + [day_names.get(d, f"周{d}") for d in range(1, 6)]
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            current_row += 1

            # 数据行
            for period in range(1, max_period + 1):
                worksheet.cell(row=current_row, column=1, value=f"第{period}节").border = thin_border
                worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")

                for day in range(1, 6):
                    # 找该时间段的所有课程
                    cell_entries = [
                        e for e in group_entries
                        if e.get("day_of_week") == day
                        and e.get("period_start") <= period <= e.get("period_end", e.get("period_start", 1))
                    ]
                    cell_text = "\n".join([
                        f"{e.get('course_name', '')}\n{e.get('teacher_name', '')}\n{e.get('classroom_name', '')}"
                        for e in cell_entries
                    ])
                    cell = worksheet.cell(row=current_row, column=day + 1, value=cell_text)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border
                    if period % 2 == 0:
                        cell.fill = alt_fill

                current_row += 1

            current_row += 1  # 分组间空行

        # 调整列宽
        worksheet.column_dimensions['A'].width = 10
        for col_idx in range(2, 7):
            from openpyxl.utils import get_column_letter
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 22

    @staticmethod
    def export_to_pdf(schedule_entries: List[Dict], metadata: Dict = None,
                      view_type: str = "class") -> bytes:
        """导出排课方案到PDF格式

        Args:
            schedule_entries: 排课条目
            metadata: 元数据
            view_type: 视图类型 (class/teacher/classroom)

        Returns:
            PDF文件的二进制内容
        """
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, PageBreak)
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        output = io.BytesIO()

        # 使用横向A4纸
        doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                               rightMargin=15*mm, leftMargin=15*mm,
                               topMargin=15*mm, bottomMargin=15*mm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=16, spaceAfter=12, alignment=1)

        elements = []

        # 标题
        title = f"排课方案 - {view_type}视图"
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 10*mm))

        # 元数据
        if metadata:
            meta_text = (f"版本: {metadata.get('version', '')} | "
                        f"质量评分: {metadata.get('quality_score', 0)}分 | "
                        f"冲突: 硬{metadata.get('hard_conflicts', 0)}/软{metadata.get('soft_conflicts', 0)}")
            elements.append(Paragraph(meta_text, styles['Normal']))
            elements.append(Spacer(1, 5*mm))

        # 按视图分组
        group_field = {"class": "class_name", "teacher": "teacher_name",
                      "classroom": "classroom_name"}.get(view_type, "class_name")

        groups: Dict[str, List[Dict]] = {}
        for entry in schedule_entries:
            key = entry.get(group_field, "未分类")
            if key not in groups:
                groups[key] = []
            groups[key].append(entry)

        day_names = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五"}

        for group_name, group_entries in sorted(groups.items()):
            # 分组标题
            elements.append(Paragraph(f"<b>{group_name}</b>", styles['Heading3']))

            # 构建表格
            table_data = [["节次"] + [day_names.get(d, f"") for d in range(1, 6)]]

            for period in range(1, 9):
                row = [f"第{period}节"]
                for day in range(1, 6):
                    cell_entries = [
                        e for e in group_entries
                        if e.get("day_of_week") == day
                        and e.get("period_start") <= period <= e.get("period_end", e.get("period_start", 1))
                    ]
                    cell_text = "\n".join([
                        f"{e.get('course_name', '')} | {e.get('classroom_name', '')}"
                        for e in cell_entries
                    ]) if cell_entries else ""
                    row.append(cell_text)
                table_data.append(row)

            col_widths = [50] + [130] * 5
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#D9E2F3')]),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 5*mm))

        doc.build(elements)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_analysis_report(schedule_result: Dict, conflicts: Dict,
                                 explanation: str, learning_data: Dict = None) -> Dict[str, Any]:
        """生成智能分析报告

        Returns:
            包含各项分析指标的报告字典
        """
        entries = schedule_result.get("entries", [])

        # 统计分析
        total_entries = len(entries)
        courses_count = len(set(e.get("course_id") for e in entries))
        teachers_count = len(set(e.get("teacher_id") for e in entries))
        classes_count = len(set(e.get("class_id") for e in entries))
        rooms_count = len(set(e.get("classroom_id") for e in entries))

        # 每日课时分布
        day_distribution = {}
        for day in range(1, 6):
            day_entries = [e for e in entries if e.get("day_of_week") == day]
            day_distribution[f"周{day}"] = len(day_entries)

        # 教室利用率(假设每天8节课)
        room_utilization = {}
        for room_id in set(e.get("classroom_id") for e in entries):
            room_entries = [e for e in entries if e.get("classroom_id") == room_id]
            room_name = room_entries[0].get("classroom_name", f"教室{room_id}")
            usage = len(room_entries)
            utilization_rate = round(usage / (5 * 8) * 100, 1)  # 5天*8节=40个时段
            room_utilization[room_name] = {
                "total_slots_used": usage,
                "utilization_rate": utilization_rate,
            }

        # 教师课时分布
        teacher_workload = {}
        for teacher_id in set(e.get("teacher_id") for e in entries):
            teacher_entries = [e for e in entries if e.get("teacher_id") == teacher_id]
            teacher_name = teacher_entries[0].get("teacher_name", f"教师{teacher_id}")
            total_periods = sum(
                (e.get("period_end", e.get("period_start", 1))
                 - e.get("period_start", 1) + 1)
                for e in teacher_entries
            )
            teacher_workload[teacher_name] = total_periods

        report = {
            "report_time": datetime.now().isoformat(),
            "schedule_version": schedule_result.get("version", ""),
            "quality_score": schedule_result.get("quality_score", 0),
            "execution_time": schedule_result.get("execution_time", 0),

            # 基础统计
            "summary": {
                "total_entries": total_entries,
                "unique_courses": courses_count,
                "unique_teachers": teachers_count,
                "unique_classes": classes_count,
                "unique_classrooms": rooms_count,
                "hard_conflicts": schedule_result.get("hard_conflicts", 0),
                "soft_conflicts": schedule_result.get("soft_conflicts", 0),
            },

            # 分布分析
            "daily_distribution": day_distribution,
            "daily_balance_score": ExportService._calc_balance_score([
                day_distribution.get(f"周{d}", 0) for d in range(1, 6)
            ]),

            # 教室利用率
            "room_utilization": room_utilization,
            "avg_room_utilization": round(
                sum(r["utilization_rate"] for r in room_utilization.values())
                / max(1, len(room_utilization)), 1
            ),

            # 教师工作量
            "teacher_workload": teacher_workload,
            "avg_teacher_workload": round(
                sum(teacher_workload.values()) / max(1, len(teacher_workload)), 1
            ),

            # AI说明
            "ai_explanation": explanation,

            # 约束学习（如有）
            "constraint_learning": learning_data.get("summary", "") if learning_data else "",

            # 冲突详情
            "conflict_summary": {
                "hard_by_category": conflicts.get("conflict_stats", {}).get("hard_by_category", {}),
                "soft_by_category": conflicts.get("conflict_stats", {}).get("soft_by_category", {}),
            },
        }

        return report

    @staticmethod
    def _calc_balance_score(values: List[float]) -> float:
        """计算分布均衡度评分 (0-100，越高越均衡)"""
        import numpy as np
        if not values or max(values) == 0:
            return 100.0
        cv = np.std(values) / (np.mean(values) + 0.001)  # 变异系数
        score = max(0, min(100, 100 * (1 - cv)))
        return round(score, 1)
